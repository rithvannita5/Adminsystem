from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
import os
import json
import sqlite3
import io
from datetime import datetime, timezone, timedelta
from functools import wraps
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from flask import send_from_directory
import random
import string
import socket
import re


# ===== TIMEZONE CAMBODIA (UTC+7) =====
CAMBODIA_TZ = timezone(timedelta(hours=7))

def get_cambodia_time():
    """ទាញយកពេលវេលាបច្ចុប្បន្នតាមម៉ោងកម្ពុជា (UTC+7)"""
    return datetime.now(CAMBODIA_TZ)

def get_cambodia_time_str():
    """ទាញយកពេលវេលាបច្ចុប្បន្នជា String តាមម៉ោងកម្ពុជា"""
    return get_cambodia_time().strftime('%Y-%m-%d %H:%M:%S')

def get_cambodia_date_str():
    """ទាញយកកាលបរិច្ឆេទបច្ចុប្បន្នជា String តាមម៉ោងកម្ពុជា"""
    return get_cambodia_time().strftime('%Y-%m-%d')


# ===== DELETE FILE =====
def delete_file_if_exists(file_path):
    """លុបឯកសារប្រសិនបើមាន"""
    if file_path and os.path.exists(file_path):
        try:
            os.remove(file_path)
            return True
        except Exception as e:
            print(f"Error deleting file {file_path}: {e}")
            return False
    return False


# ===== GET LOCAL IP =====
def get_local_ip():
    """ទាញយក IP ម៉ាស៊ីនសម្រាប់ប្រើប្រាស់"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "127.0.0.1"


# ===== APP INIT =====
app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB


# ===== DATABASE SETUP =====
DATABASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'admin_system.db')

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


# ===== INIT DATABASE =====
def init_db():
    conn = get_db()
    cursor = conn.cursor()

    # ===== តារាងឯកសារចូល =====
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS income_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT NOT NULL,
            content TEXT NOT NULL,
            source TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            unit_type TEXT NOT NULL,
            price_per_page REAL NOT NULL,
            total_price REAL NOT NULL,
            village TEXT NOT NULL,
            entry_date TEXT NOT NULL,
            created_at TEXT NOT NULL,
            created_by TEXT NOT NULL
        )
    ''')

    # ===== តារាងឯកសារចេញ =====
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS expense_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT NOT NULL,
            content TEXT NOT NULL,
            source TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            unit_type TEXT NOT NULL,
            price_per_page REAL NOT NULL,
            total_price REAL NOT NULL,
            recipient TEXT NOT NULL,
            entry_date TEXT NOT NULL,
            created_at TEXT NOT NULL,
            created_by TEXT NOT NULL
        )
    ''')

    # ===== តារាងអ្នកប្រើប្រាស់ =====
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            full_name TEXT,
            phone TEXT,
            role_id INTEGER DEFAULT 2,
            is_active INTEGER DEFAULT 1,
            created_at TEXT NOT NULL,
            created_by TEXT NOT NULL
        )
    ''')

    # ===== តារាងតួនាទី =====
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS roles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT,
            permissions TEXT,
            created_at TEXT NOT NULL,
            created_by TEXT NOT NULL
        )
    ''')

    # ===== តារាងសវនកម្ម =====
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS audit_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            username TEXT,
            action TEXT NOT NULL,
            module TEXT NOT NULL,
            record_id INTEGER,
            details TEXT,
            ip_address TEXT,
            created_at TEXT NOT NULL
        )
    ''')

    # ===== តារាងឯកសារពត៌មាន =====
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS info_documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT NOT NULL,
            file_path TEXT NOT NULL,
            file_name TEXT NOT NULL,
            uploaded_at TEXT NOT NULL,
            uploaded_by TEXT NOT NULL
        )
    ''')

    # ===== តារាងបុគ្គលិក =====
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            photo TEXT,
            name TEXT NOT NULL,
            gender TEXT NOT NULL,
            birth_date TEXT NOT NULL,
            birth_place TEXT NOT NULL,
            id_card TEXT NOT NULL,
            role TEXT NOT NULL,
            salary REAL NOT NULL,
            start_date TEXT NOT NULL,
            start_file TEXT,
            end_date TEXT,
            end_file TEXT,
            file_path TEXT,
            created_at TEXT NOT NULL,
            created_by TEXT NOT NULL
        )
    ''')

    # ===== តារាង OTP =====
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS otp_codes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            phone TEXT NOT NULL,
            otp_code TEXT NOT NULL,
            is_used INTEGER DEFAULT 0,
            expires_at TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    ''')

    # ===== បន្ថែមតួនាទីលំនាំដើម =====
    cursor.execute('SELECT COUNT(*) FROM roles')
    if cursor.fetchone()[0] == 0:
        cursor.execute('''
            INSERT INTO roles (name, description, permissions, created_at, created_by)
            VALUES (?, ?, ?, ?, ?)
        ''', ('Admin', 'អ្នកគ្រប់គ្រងប្រព័ន្ធ', 'all', get_cambodia_time_str(), 'system'))

        cursor.execute('''
            INSERT INTO roles (name, description, permissions, created_at, created_by)
            VALUES (?, ?, ?, ?, ?)
        ''', ('User', 'អ្នកប្រើប្រាស់ធម្មតា',
        'dashboard,income_bank_view,income_land_view,income_home_view,income_wedding_view,income_other_view,expense_view,income_budget_view,expense_budget_view,total_budget_view,customers_view,info_view,settings_view,employees_view',
        get_cambodia_time_str(), 'system'))

    # ===== បន្ថែមអ្នកប្រើប្រាស់លំនាំដើម =====
    cursor.execute('SELECT COUNT(*) FROM users')
    if cursor.fetchone()[0] == 0:
        cursor.execute('''
            INSERT INTO users (username, password, full_name, phone, role_id, is_active, created_at, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', ('admin', 'password123', 'អ្នកគ្រប់គ្រង', '012345678', 1, 1, get_cambodia_time_str(), 'system'))

        cursor.execute('''
            INSERT INTO users (username, password, full_name, phone, role_id, is_active, created_at, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', ('user', 'userpass', 'អ្នកប្រើប្រាស់', '098765432', 2, 1, get_cambodia_time_str(), 'system'))

        cursor.execute('''
            INSERT INTO users (username, password, full_name, phone, role_id, is_active, created_at, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', ('sok', 'password123', 'សុខ សុភាព', '097654321', 2, 1, get_cambodia_time_str(), 'system'))

    conn.commit()
    conn.close()


# ===== LOGIN REQUIRED DECORATOR =====
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


# ===== PERMISSION CHECK =====
PERMISSIONS_LIST = {
    'dashboard': 'មើល Dashboard',

    # ===== ឯកសារចូល =====
    'income_bank_view': 'មើលការងារធនាគារ',
    'income_bank_add': 'បន្ថែមការងារធនាគារ',
    'income_bank_edit': 'កែប្រែការងារធនាគារ',
    'income_bank_delete': 'លុបការងារធនាគារ',
    'income_land_view': 'មើលការងារដីធ្លី',
    'income_land_add': 'បន្ថែមការងារដីធ្លី',
    'income_land_edit': 'កែប្រែការងារដីធ្លី',
    'income_land_delete': 'លុបការងារដីធ្លី',
    'income_home_view': 'មើលការងារផ្ទះ',
    'income_home_add': 'បន្ថែមការងារផ្ទះ',
    'income_home_edit': 'កែប្រែការងារផ្ទះ',
    'income_home_delete': 'លុបការងារផ្ទះ',
    'income_wedding_view': 'មើលការងារអាពាហ៍ពិពាហ៍',
    'income_wedding_add': 'បន្ថែមការងារអាពាហ៍ពិពាហ៍',
    'income_wedding_edit': 'កែប្រែការងារអាពាហ៍ពិពាហ៍',
    'income_wedding_delete': 'លុបការងារអាពាហ៍ពិពាហ៍',
    'income_other_view': 'មើលការងារផ្សេងៗ',
    'income_other_add': 'បន្ថែមការងារផ្សេងៗ',
    'income_other_edit': 'កែប្រែការងារផ្សេងៗ',
    'income_other_delete': 'លុបការងារផ្សេងៗ',

    # ===== ឯកសារចេញ =====
    'expense_report_view': 'មើលការងាររបាយការណ៍',
    'expense_report_add': 'បន្ថែមការងាររបាយការណ៍',
    'expense_report_edit': 'កែប្រែការងាររបាយការណ៍',
    'expense_report_delete': 'លុបការងាររបាយការណ៍',
    'expense_decision_view': 'មើលការងារសេចក្តីសម្រេច',
    'expense_decision_add': 'បន្ថែមការងារសេចក្តីសម្រេច',
    'expense_decision_edit': 'កែប្រែការងារសេចក្តីសម្រេច',
    'expense_decision_delete': 'លុបការងារសេចក្តីសម្រេច',
    'expense_warrant_view': 'មើលការងារចេញដីការ',
    'expense_warrant_add': 'បន្ថែមការងារចេញដីការ',
    'expense_warrant_edit': 'កែប្រែការងារចេញដីការ',
    'expense_warrant_delete': 'លុបការងារចេញដីការ',
    'expense_financial_view': 'មើលការងារហិរញ្ញវត្ថុ',
    'expense_financial_add': 'បន្ថែមការងារហិរញ្ញវត្ថុ',
    'expense_financial_edit': 'កែប្រែការងារហិរញ្ញវត្ថុ',
    'expense_financial_delete': 'លុបការងារហិរញ្ញវត្ថុ',
    'expense_other_view': 'មើលការងារឯកសារចេញផ្សេងៗ',
    'expense_other_add': 'បន្ថែមការងារឯកសារចេញផ្សេងៗ',
    'expense_other_edit': 'កែប្រែការងារឯកសារចេញផ្សេងៗ',
    'expense_other_delete': 'លុបការងារឯកសារចេញផ្សេងៗ',

    # ===== ថវិកា =====
    'income_budget_view': 'មើលថវិកាចូល',
    'expense_budget_view': 'មើលថវិកាចេញ',
    'total_budget_view': 'មើលថវិការសរុប',

    # ===== ផ្សេងៗ =====
    'customers_view': 'មើលរបកអតិថិជន',
    'info_view': 'មើលពត៌មាន',
    'info_upload': 'ផ្ទុកឯកសារពត៌មាន',
    'employees_view': 'មើលបុគ្គលិក',
    'employees_add': 'បន្ថែមបុគ្គលិក',
    'employees_edit': 'កែប្រែបុគ្គលិក',
    'employees_delete': 'លុបបុគ្គលិក',
    'settings_view': 'មើលការកំណត់',
    'settings_users': 'គ្រប់គ្រងអ្នកប្រើប្រាស់',
    'settings_roles': 'គ្រប់គ្រងតួនាទី',
    'settings_audit': 'មើលសវនកម្ម',
}

def has_permission(permission):
    if 'role_id' not in session:
        return False

    if session.get('role_id') == 1:
        return True

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT permissions FROM roles WHERE id = ?', (session.get('role_id'),))
    row = cursor.fetchone()
    conn.close()

    if not row or not row['permissions']:
        return False

    permissions = row['permissions'].split(',')
    permissions = [p.strip() for p in permissions]

    return permission in permissions or 'all' in permissions

def get_user_permissions():
    if 'role_id' not in session:
        return []

    if session.get('role_id') == 1:
        return ['all']

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT permissions FROM roles WHERE id = ?', (session.get('role_id'),))
    row = cursor.fetchone()
    conn.close()

    if not row or not row['permissions']:
        return []

    permissions = row['permissions'].split(',')
    return [p.strip() for p in permissions]


# ============================================
# AUDIT LOG FUNCTION
# ============================================
def log_audit(action, module, record_id=None, details=None):
    """កត់ត្រាសកម្មភាពអ្នកប្រើប្រាស់"""
    try:
        conn = get_db()
        cursor = conn.cursor()

        ip_address = request.headers.get('X-Forwarded-For', request.remote_addr)

        cursor.execute('''
            INSERT INTO audit_logs (user_id, username, action, module, record_id, details, ip_address, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            session.get('user_id'),
            session.get('username', 'unknown'),
            action,
            module,
            record_id,
            details,
            ip_address,
            get_cambodia_time_str()
        ))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error logging audit: {e}")


# ============================================
# GENERATE RANDOM PASSWORD
# ============================================
def generate_random_password(length=10):
    """បង្កើតពាក្យសម្ងាត់ថ្មីដោយស្វ័យប្រវត្តិ"""
    characters = string.ascii_letters + string.digits + '!@#$%^&*'
    return ''.join(random.choice(characters) for i in range(length))


# ============================================
# GENERATE OTP
# ============================================
def generate_otp(length=6):
    """បង្កើត OTP ចៃដន្យ"""
    return ''.join(random.choices('0123456789', k=length))


# ===== JINJA2 CONTEXT PROCESSOR =====
@app.context_processor
def utility_processor():
    def get_user_permissions():
        if 'role_id' not in session:
            return []

        if session.get('role_id') == 1:
            return ['all']

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT permissions FROM roles WHERE id = ?', (session.get('role_id'),))
        row = cursor.fetchone()
        conn.close()

        if not row or not row['permissions']:
            return []

        permissions = row['permissions'].split(',')
        return [p.strip() for p in permissions]

    def get_version():
        return '20260814-dmMM'

    return dict(get_user_permissions=get_user_permissions, get_version=get_version)


# ============================================
# ROUTE: INSTALL (PWA)
# ============================================
@app.route('/install')
def install():
    return render_template('install.html')


# ============================================
# ROUTE: OFFLINE PAGE
# ============================================
@app.route('/offline')
def offline():
    return '''
    <!DOCTYPE html>
    <html lang="km">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Offline - គ្មានការតភ្ជាប់</title>
        <link href="https://fonts.googleapis.com/css2?family=Khmer&display=swap" rel="stylesheet">
        <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.0/css/all.min.css">
        <style>
            body {
                font-family: 'Khmer', 'Khmer OS', sans-serif;
                background: #1a1a2e;
                display: flex;
                justify-content: center;
                align-items: center;
                min-height: 100vh;
                margin: 0;
                color: white;
                text-align: center;
                padding: 20px;
            }
            .offline-box {
                background: rgba(255,255,255,0.05);
                backdrop-filter: blur(10px);
                padding: 50px 40px;
                border-radius: 20px;
                border: 1px solid rgba(255,255,255,0.08);
                max-width: 400px;
            }
            .offline-box i {
                font-size: 64px;
                color: #FFD700;
                margin-bottom: 20px;
            }
            .offline-box h1 {
                font-size: 24px;
                margin-bottom: 10px;
            }
            .offline-box p {
                color: rgba(255,255,255,0.5);
                font-size: 14px;
                line-height: 1.6;
            }
            .offline-box .btn {
                display: inline-block;
                margin-top: 20px;
                padding: 12px 30px;
                background: #FFD700;
                color: #1a1a2e;
                text-decoration: none;
                border-radius: 10px;
                font-weight: bold;
                font-family: 'Khmer', 'Khmer OS', sans-serif;
            }
            .offline-box .btn:hover {
                background: #f0c000;
            }
        </style>
    </head>
    <body>
        <div class="offline-box">
            <i class="fas fa-wifi-slash"></i>
            <h1>📡 គ្មានការតភ្ជាប់</h1>
            <p>សូមពិនិត្យមើលការតភ្ជាប់អ៊ីនធឺណិតរបស់អ្នក<br>
            ទិន្នន័យនឹងត្រូវបានរក្សាទុកក្នុងស្រុក និង Sync ស្វ័យប្រវត្តិពេលមានបណ្តាញ</p>
            <a href="/dashboard" class="btn"><i class="fas fa-arrow-left"></i> ត្រឡប់ទៅ Dashboard</a>
        </div>
    </body>
    </html>
    '''


# ============================================
# ROUTE: PWA Manifest
# ============================================
@app.route('/static/manifest.json')
def serve_manifest():
    try:
        return send_file('static/manifest.json', mimetype='application/json')
    except:
        return jsonify({'error': 'Manifest not found'}), 404


# ============================================
# ROUTE: Service Worker
# ============================================
@app.route('/static/service-worker.js')
def serve_service_worker():
    try:
        return send_file('static/service-worker.js', mimetype='application/javascript')
    except:
        return jsonify({'error': 'Service Worker not found'}), 404


# ============================================
# ROUTE: ទំព័រដើម
# ============================================
@app.route('/')
def home():
    if 'username' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


# ============================================
# ROUTE: Login
# ============================================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM users WHERE username = ? AND password = ? AND is_active = 1', (username, password))
        user = cursor.fetchone()
        conn.close()

        if user:
            session['username'] = username
            session['user_id'] = user['id']
            session['role_id'] = user['role_id']
            session['login_time'] = get_cambodia_time_str()

            log_audit(
                action='ចូលប្រើប្រាស់',
                module='ប្រព័ន្ធ',
                details=f"អ្នកប្រើប្រាស់ {username} បានចូលប្រើប្រាស់ប្រព័ន្ធ"
            )

            flash('ចូលប្រើប្រាស់ដោយជោគជ័យ!', 'success')
            return redirect(url_for('dashboard'))
        else:
            error = 'ឈ្មោះអ្នកប្រើប្រាស់ ឬ ពាក្យសម្ងាត់មិនត្រឹមត្រូវ!'
            return render_template('login.html', error=error)

    return render_template('login.html')


# ============================================
# ROUTE: Logout
# ============================================
@app.route('/logout')
def logout():
    username = session.get('username', 'unknown')

    log_audit(
        action='ចាកចេញ',
        module='ប្រព័ន្ធ',
        details=f"អ្នកប្រើប្រាស់ {username} បានចាកចេញពីប្រព័ន្ធ"
    )

    session.pop('username', None)
    session.pop('user_id', None)
    session.pop('role_id', None)
    session.pop('login_time', None)
    flash('អ្នកបានចាកចេញពីប្រព័ន្ធ', 'info')
    return redirect(url_for('login'))


# ============================================
# ROUTE: Dashboard
# ============================================
@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT SUM(total_price) FROM income_records')
    total_income = cursor.fetchone()[0] or 0

    cursor.execute('SELECT SUM(total_price) FROM expense_records')
    total_expense = cursor.fetchone()[0] or 0

    cursor.execute('SELECT COUNT(*) FROM income_records')
    income_count = cursor.fetchone()[0] or 0

    cursor.execute('SELECT COUNT(*) FROM expense_records')
    expense_count = cursor.fetchone()[0] or 0

    cursor.execute('SELECT COUNT(DISTINCT village) FROM income_records')
    customers_from_income = cursor.fetchone()[0] or 0

    cursor.execute('SELECT COUNT(DISTINCT recipient) FROM expense_records')
    customers_from_expense = cursor.fetchone()[0] or 0

    total_customers = customers_from_income + customers_from_expense

    income_types = ['bank', 'land', 'home', 'wedding', 'other']
    income_labels = {
        'bank': 'ធនាគារ',
        'land': 'ដីធ្លី',
        'home': 'ផ្ទះ',
        'wedding': 'អាពាហ៍ពិពាហ៍',
        'other': 'ផ្សេងៗ'
    }
    income_colors = {
        'bank': '#4CAF50',
        'land': '#2196F3',
        'home': '#FF9800',
        'wedding': '#E91E63',
        'other': '#9C27B0'
    }

    income_by_type = {}
    for t in income_types:
        cursor.execute('SELECT SUM(total_price) FROM income_records WHERE type = ?', (t,))
        amount = cursor.fetchone()[0] or 0
        income_by_type[t] = amount

    chart_labels = [income_labels[t] for t in income_types]
    chart_values = [income_by_type.get(t, 0) for t in income_types]
    chart_colors = [income_colors[t] for t in income_types]

    cursor.execute('SELECT * FROM income_records ORDER BY id DESC LIMIT 5')
    recent_income = cursor.fetchall()

    cursor.execute('SELECT * FROM expense_records ORDER BY id DESC LIMIT 5')
    recent_expense = cursor.fetchall()

    conn.close()

    stats = {
        'total_income': total_income,
        'total_expense': total_expense,
        'balance': total_income - total_expense,
        'total_customers': total_customers,
        'income_count': income_count,
        'expense_count': expense_count,
        'customers_from_income': customers_from_income,
        'customers_from_expense': customers_from_expense
    }

    return render_template('dashboard.html',
                         username=session['username'],
                         stats=stats,
                         income_records=recent_income,
                         expense_records=recent_expense,
                         chart_labels=chart_labels,
                         chart_values=chart_values,
                         chart_colors=chart_colors,
                         permissions=get_user_permissions())


# ============================================
# ROUTE: ឯកសារចូល (Income)
# ============================================
@app.route('/income')
@login_required
def income():
    user_permissions = get_user_permissions()
    return render_template('income.html',
                         username=session['username'],
                         permissions=user_permissions)


# ===== API: GET Income Records =====
@app.route('/api/income', methods=['GET'])
@login_required
def api_get_income():
    type_filter = request.args.get('type', 'bank')
    search = request.args.get('search', '').strip().lower()

    perm_map = {
        'bank': 'income_bank_view',
        'land': 'income_land_view',
        'home': 'income_home_view',
        'wedding': 'income_wedding_view',
        'other': 'income_other_view'
    }

    if type_filter in perm_map and not has_permission(perm_map[type_filter]):
        return jsonify([])

    conn = get_db()
    cursor = conn.cursor()

    query = 'SELECT * FROM income_records WHERE type = ?'
    params = [type_filter]

    if search:
        query += ' AND (LOWER(content) LIKE ? OR LOWER(source) LIKE ? OR LOWER(village) LIKE ?)'
        search_param = f'%{search}%'
        params.extend([search_param, search_param, search_param])

    query += ' ORDER BY id DESC'
    cursor.execute(query, params)
    records = cursor.fetchall()
    conn.close()

    result = []
    for row in records:
        result.append({
            'id': row['id'],
            'type': row['type'],
            'content': row['content'],
            'source': row['source'],
            'quantity': row['quantity'],
            'unit_type': row['unit_type'],
            'price_per_page': row['price_per_page'],
            'total_price': row['total_price'],
            'village': row['village'],
            'entry_date': row['entry_date']
        })

    return jsonify(result)


# ===== API: ADD Income Record =====
@app.route('/api/income', methods=['POST'])
@login_required
def api_add_income():
    data = request.json
    record_type = data.get('type', 'bank')

    perm_map = {
        'bank': 'income_bank_add',
        'land': 'income_land_add',
        'home': 'income_home_add',
        'wedding': 'income_wedding_add',
        'other': 'income_other_add'
    }

    if record_type in perm_map and not has_permission(perm_map[record_type]):
        return jsonify({'success': False, 'error': 'អ្នកមិនមានសិទ្ធិបន្ថែមឯកសារប្រភេទនេះទេ!'})

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO income_records
        (type, content, source, quantity, unit_type, price_per_page, total_price, village, entry_date, created_at, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        record_type,
        data.get('content', ''),
        data.get('source', ''),
        data.get('quantity', 0),
        data.get('unit_type', 'ច្បាប់'),
        data.get('price_per_page', 0),
        data.get('total_price', 0),
        data.get('village', ''),
        data.get('entry_date', ''),
        get_cambodia_time_str(),
        session['username']
    ))
    conn.commit()
    new_id = cursor.lastrowid
    conn.close()

    log_audit(
        action='បន្ថែម',
        module='ឯកសារចូល',
        record_id=new_id,
        details=f"បានបន្ថែមឯកសារចូលប្រភេទ {record_type}: {data.get('content', '')[:50]}..."
    )

    return jsonify({'success': True, 'id': new_id})


# ===== API: UPDATE Income Record =====
@app.route('/api/income/<int:record_id>', methods=['PUT'])
@login_required
def api_update_income(record_id):
    data = request.json
    record_type = data.get('type', 'bank')

    perm_map = {
        'bank': 'income_bank_edit',
        'land': 'income_land_edit',
        'home': 'income_home_edit',
        'wedding': 'income_wedding_edit',
        'other': 'income_other_edit'
    }

    if record_type in perm_map and not has_permission(perm_map[record_type]):
        return jsonify({'success': False, 'error': 'អ្នកមិនមានសិទ្ធិកែប្រែឯកសារប្រភេទនេះទេ!'})

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE income_records
        SET content = ?, source = ?, quantity = ?, unit_type = ?,
            price_per_page = ?, total_price = ?, village = ?, entry_date = ?
        WHERE id = ?
    ''', (
        data.get('content', ''),
        data.get('source', ''),
        data.get('quantity', 0),
        data.get('unit_type', 'ច្បាប់'),
        data.get('price_per_page', 0),
        data.get('total_price', 0),
        data.get('village', ''),
        data.get('entry_date', ''),
        record_id
    ))
    conn.commit()
    conn.close()

    log_audit(
        action='កែប្រែ',
        module='ឯកសារចូល',
        record_id=record_id,
        details=f"បានកែប្រែឯកសារចូល ID {record_id}"
    )

    return jsonify({'success': True})


# ===== API: DELETE Income Record =====
@app.route('/api/income/<int:record_id>', methods=['DELETE'])
@login_required
def api_delete_income(record_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT type, content FROM income_records WHERE id = ?', (record_id,))
    row = cursor.fetchone()

    if row:
        record_type = row['type']
        perm_map = {
            'bank': 'income_bank_delete',
            'land': 'income_land_delete',
            'home': 'income_home_delete',
            'wedding': 'income_wedding_delete',
            'other': 'income_other_delete'
        }

        if record_type in perm_map and not has_permission(perm_map[record_type]):
            conn.close()
            return jsonify({'success': False, 'error': 'អ្នកមិនមានសិទ្ធិលុបឯកសារប្រភេទនេះទេ!'})

    cursor.execute('DELETE FROM income_records WHERE id = ?', (record_id,))
    conn.commit()
    conn.close()

    log_audit(
        action='លុប',
        module='ឯកសារចូល',
        record_id=record_id,
        details=f"បានលុបឯកសារចូល ID {record_id}"
    )

    return jsonify({'success': True})


# ===== API: EXPORT Income to Excel =====
@app.route('/api/income/export', methods=['GET'])
@login_required
def api_export_income():
    type_filter = request.args.get('type', 'bank')
    search = request.args.get('search', '').strip().lower()

    conn = get_db()
    cursor = conn.cursor()

    query = 'SELECT * FROM income_records WHERE type = ?'
    params = [type_filter]

    if search:
        query += ' AND (LOWER(content) LIKE ? OR LOWER(source) LIKE ? OR LOWER(village) LIKE ?)'
        search_param = f'%{search}%'
        params.extend([search_param, search_param, search_param])

    query += ' ORDER BY id DESC'
    cursor.execute(query, params)
    records = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ឯកសារចូល"

    type_names = {
        'bank': 'ធនាគារ',
        'land': 'ដីធ្លី',
        'home': 'ផ្ទះ',
        'wedding': 'អាពាហ៍ពិពាហ៍',
        'other': 'ផ្សេងៗ'
    }
    type_display = type_names.get(type_filter, type_filter)

    headers = ['លរ', 'ខ្លឹមសារលិខិត', 'ប្រភពលិខិត', 'ចំនួន', 'ច្បាប់/សន្លឹក', 'តម្លៃ/១ច្បាប់', 'តម្លៃសេវាសរុប', 'មកពីភូមិ', 'ថ្ងៃខែចូល']
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for idx, row in enumerate(records, 1):
        ws.append([
            idx,
            row['content'],
            row['source'],
            row['quantity'],
            row['unit_type'],
            row['price_per_page'],
            row['total_price'],
            row['village'],
            row['entry_date']
        ])

    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"ឯកសារចូល_{type_display}_{get_cambodia_date_str()}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ============================================
# ROUTE: ឯកសារចេញ (Expense)
# ============================================
@app.route('/expense')
@login_required
def expense():
    user_permissions = get_user_permissions()
    return render_template('expense.html',
                         username=session['username'],
                         permissions=user_permissions)


# ===== API: GET Expense Records =====
@app.route('/api/expense', methods=['GET'])
@login_required
def api_get_expense():
    type_filter = request.args.get('type', 'report')
    search = request.args.get('search', '').strip().lower()

    perm_map = {
        'report': 'expense_report_view',
        'decision': 'expense_decision_view',
        'warrant': 'expense_warrant_view',
        'financial': 'expense_financial_view',
        'other': 'expense_other_view'
    }

    if type_filter in perm_map and not has_permission(perm_map[type_filter]):
        return jsonify([])

    conn = get_db()
    cursor = conn.cursor()

    query = 'SELECT * FROM expense_records WHERE type = ?'
    params = [type_filter]

    if search:
        query += ' AND (LOWER(content) LIKE ? OR LOWER(source) LIKE ? OR LOWER(recipient) LIKE ?)'
        search_param = f'%{search}%'
        params.extend([search_param, search_param, search_param])

    query += ' ORDER BY id DESC'
    cursor.execute(query, params)
    records = cursor.fetchall()
    conn.close()

    result = []
    for row in records:
        result.append({
            'id': row['id'],
            'type': row['type'],
            'content': row['content'],
            'source': row['source'],
            'quantity': row['quantity'],
            'unit_type': row['unit_type'],
            'price_per_page': row['price_per_page'],
            'total_price': row['total_price'],
            'recipient': row['recipient'],
            'entry_date': row['entry_date']
        })

    return jsonify(result)


# ===== API: ADD Expense Record =====
@app.route('/api/expense', methods=['POST'])
@login_required
def api_add_expense():
    data = request.json
    record_type = data.get('type', 'report')

    perm_map = {
        'report': 'expense_report_add',
        'decision': 'expense_decision_add',
        'warrant': 'expense_warrant_add',
        'financial': 'expense_financial_add',
        'other': 'expense_other_add'
    }

    if record_type in perm_map and not has_permission(perm_map[record_type]):
        return jsonify({'success': False, 'error': 'អ្នកមិនមានសិទ្ធិបន្ថែមឯកសារប្រភេទនេះទេ!'})

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO expense_records
        (type, content, source, quantity, unit_type, price_per_page, total_price, recipient, entry_date, created_at, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        record_type,
        data.get('content', ''),
        data.get('source', ''),
        data.get('quantity', 0),
        data.get('unit_type', 'ច្បាប់'),
        data.get('price_per_page', 0),
        data.get('total_price', 0),
        data.get('recipient', ''),
        data.get('entry_date', ''),
        get_cambodia_time_str(),
        session['username']
    ))
    conn.commit()
    new_id = cursor.lastrowid
    conn.close()

    log_audit(
        action='បន្ថែម',
        module='ឯកសារចេញ',
        record_id=new_id,
        details=f"បានបន្ថែមឯកសារចេញប្រភេទ {record_type}: {data.get('content', '')[:50]}..."
    )

    return jsonify({'success': True, 'id': new_id})


# ===== API: UPDATE Expense Record =====
@app.route('/api/expense/<int:record_id>', methods=['PUT'])
@login_required
def api_update_expense(record_id):
    data = request.json
    record_type = data.get('type', 'report')

    perm_map = {
        'report': 'expense_report_edit',
        'decision': 'expense_decision_edit',
        'warrant': 'expense_warrant_edit',
        'financial': 'expense_financial_edit',
        'other': 'expense_other_edit'
    }

    if record_type in perm_map and not has_permission(perm_map[record_type]):
        return jsonify({'success': False, 'error': 'អ្នកមិនមានសិទ្ធិកែប្រែឯកសារប្រភេទនេះទេ!'})

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE expense_records
        SET content = ?, source = ?, quantity = ?, unit_type = ?,
            price_per_page = ?, total_price = ?, recipient = ?, entry_date = ?
        WHERE id = ?
    ''', (
        data.get('content', ''),
        data.get('source', ''),
        data.get('quantity', 0),
        data.get('unit_type', 'ច្បាប់'),
        data.get('price_per_page', 0),
        data.get('total_price', 0),
        data.get('recipient', ''),
        data.get('entry_date', ''),
        record_id
    ))
    conn.commit()
    conn.close()

    log_audit(
        action='កែប្រែ',
        module='ឯកសារចេញ',
        record_id=record_id,
        details=f"បានកែប្រែឯកសារចេញ ID {record_id}"
    )

    return jsonify({'success': True})


# ===== API: DELETE Expense Record =====
@app.route('/api/expense/<int:record_id>', methods=['DELETE'])
@login_required
def api_delete_expense(record_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT type, content FROM expense_records WHERE id = ?', (record_id,))
    row = cursor.fetchone()

    if row:
        record_type = row['type']
        perm_map = {
            'report': 'expense_report_delete',
            'decision': 'expense_decision_delete',
            'warrant': 'expense_warrant_delete',
            'financial': 'expense_financial_delete',
            'other': 'expense_other_delete'
        }

        if record_type in perm_map and not has_permission(perm_map[record_type]):
            conn.close()
            return jsonify({'success': False, 'error': 'អ្នកមិនមានសិទ្ធិលុបឯកសារប្រភេទនេះទេ!'})

    cursor.execute('DELETE FROM expense_records WHERE id = ?', (record_id,))
    conn.commit()
    conn.close()

    log_audit(
        action='លុប',
        module='ឯកសារចេញ',
        record_id=record_id,
        details=f"បានលុបឯកសារចេញ ID {record_id}"
    )

    return jsonify({'success': True})


# ===== API: EXPORT Expense to Excel =====
@app.route('/api/expense/export', methods=['GET'])
@login_required
def api_export_expense():
    type_filter = request.args.get('type', 'report')
    search = request.args.get('search', '').strip().lower()

    conn = get_db()
    cursor = conn.cursor()

    query = 'SELECT * FROM expense_records WHERE type = ?'
    params = [type_filter]

    if search:
        query += ' AND (LOWER(content) LIKE ? OR LOWER(source) LIKE ? OR LOWER(recipient) LIKE ?)'
        search_param = f'%{search}%'
        params.extend([search_param, search_param, search_param])

    query += ' ORDER BY id DESC'
    cursor.execute(query, params)
    records = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ឯកសារចេញ"

    type_names = {
        'report': 'របាយការណ៍',
        'decision': 'សេចក្តីសម្រេច',
        'warrant': 'ដីការ',
        'financial': 'ហិរញ្ញវត្ថុ',
        'other': 'ផ្សេងៗ'
    }
    type_display = type_names.get(type_filter, type_filter)

    headers = ['លរ', 'ខ្លឹមសារលិខិត', 'ប្រភពលិខិត', 'ចំនួន', 'ច្បាប់/សន្លឹក', 'តម្លៃ/១ច្បាប់', 'តម្លៃសេវាសរុប', 'អង្គភាពទទួល', 'ថ្ងៃខែចូល']
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for idx, row in enumerate(records, 1):
        ws.append([
            idx,
            row['content'],
            row['source'],
            row['quantity'],
            row['unit_type'],
            row['price_per_page'],
            row['total_price'],
            row['recipient'],
            row['entry_date']
        ])

    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"ឯកសារចេញ_{type_display}_{get_cambodia_date_str()}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ============================================
# ROUTE: ថវិកាឯកសារចូល (Income Budget)
# ============================================
@app.route('/income-budget')
@login_required
def income_budget():
    conn = get_db()
    cursor = conn.cursor()

    income_types = ['bank', 'land', 'home', 'wedding', 'other']
    income_labels = {
        'bank': 'ធនាគារ',
        'land': 'ការងារដីធ្លី',
        'home': 'ការងារផ្ទះ',
        'wedding': 'អាពាហ៍ពិពាហ៍',
        'other': 'ចូលផ្សេងៗ'
    }
    income_colors = {
        'bank': '#4CAF50',
        'land': '#2196F3',
        'home': '#FF9800',
        'wedding': '#E91E63',
        'other': '#9C27B0'
    }

    income_data = {}
    total_income_all = 0
    records_count = {}

    for t in income_types:
        cursor.execute('SELECT SUM(total_price), COUNT(*) FROM income_records WHERE type = ?', (t,))
        row = cursor.fetchone()
        amount = row[0] or 0
        count = row[1] or 0
        income_data[t] = amount
        records_count[t] = count
        total_income_all += amount

    conn.close()

    chart_labels = [income_labels[t] for t in income_types]
    chart_values = [income_data.get(t, 0) for t in income_types]
    chart_colors = [income_colors[t] for t in income_types]

    return render_template('income_budget.html',
                         username=session['username'],
                         income_data=income_data,
                         income_labels=income_labels,
                         income_colors=income_colors,
                         records_count=records_count,
                         total_income=total_income_all,
                         chart_labels=chart_labels,
                         chart_values=chart_values,
                         chart_colors=chart_colors,
                         permissions=get_user_permissions())


# ============================================
# ROUTE: របកអតិថិជនចូល (Customers)
# ============================================
@app.route('/customers')
@login_required
def customers():
    conn = get_db()
    cursor = conn.cursor()

    income_types = ['bank', 'land', 'home', 'wedding', 'other']
    income_labels = {
        'bank': 'ធនាគារ',
        'land': 'ការងារដី',
        'home': 'ការងារផ្ទះ',
        'wedding': 'អាពាហ៍ពិពាហ៍',
        'other': 'ចូលផ្សេងៗ'
    }
    customer_colors = {
        'bank': '#4CAF50',
        'land': '#2196F3',
        'home': '#FF9800',
        'wedding': '#E91E63',
        'other': '#9C27B0'
    }

    customer_data = {}
    total_customers_all = 0

    for t in income_types:
        cursor.execute('SELECT COUNT(DISTINCT village) FROM income_records WHERE type = ?', (t,))
        count = cursor.fetchone()[0] or 0
        customer_data[t] = count
        total_customers_all += count

    conn.close()

    chart_labels = [income_labels[t] for t in income_types]
    chart_values = [customer_data.get(t, 0) for t in income_types]
    chart_colors = [customer_colors[t] for t in income_types]

    return render_template('customers.html',
                         username=session['username'],
                         customer_data=customer_data,
                         customer_labels=income_labels,
                         customer_colors=customer_colors,
                         total_customers=total_customers_all,
                         chart_labels=chart_labels,
                         chart_values=chart_values,
                         chart_colors=chart_colors,
                         permissions=get_user_permissions())


# ============================================
# ROUTE: ថវិកាឯកសារចេញ (Expense Budget)
# ============================================
@app.route('/expense-budget')
@login_required
def expense_budget():
    conn = get_db()
    cursor = conn.cursor()

    expense_types = ['report', 'decision', 'warrant', 'financial', 'other']
    expense_labels = {
        'report': 'របាយការណ៍',
        'decision': 'សេចក្តីសម្រេច',
        'warrant': 'របក ដីការ',
        'financial': 'បៀរវត្ស/ហិរញ្ញវត្ថុ',
        'other': 'ឯកសារចេញផ្សេងៗ'
    }
    expense_colors = {
        'report': '#FF6B6B',
        'decision': '#FF9F43',
        'warrant': '#FECA57',
        'financial': '#48DBFB',
        'other': '#A29BFE'
    }

    expense_data = {}
    total_expense_all = 0
    people_count = {}
    total_people_all = 0

    for t in expense_types:
        cursor.execute('SELECT SUM(total_price) FROM expense_records WHERE type = ?', (t,))
        amount = cursor.fetchone()[0] or 0
        expense_data[t] = amount
        total_expense_all += amount

        cursor.execute('SELECT COUNT(DISTINCT recipient) FROM expense_records WHERE type = ?', (t,))
        count = cursor.fetchone()[0] or 0
        people_count[t] = count
        total_people_all += count

    conn.close()

    chart_labels = [expense_labels[t] for t in expense_types]
    chart_values = [people_count.get(t, 0) for t in expense_types]
    chart_colors = [expense_colors[t] for t in expense_types]

    return render_template('expense_budget.html',
                         username=session['username'],
                         expense_data=expense_data,
                         expense_labels=expense_labels,
                         expense_colors=expense_colors,
                         people_count=people_count,
                         total_people=total_people_all,
                         total_expense=total_expense_all,
                         chart_labels=chart_labels,
                         chart_values=chart_values,
                         chart_colors=chart_colors,
                         permissions=get_user_permissions())


# ============================================
# ROUTE: ថវិការសរុប (Total Budget)
# ============================================
@app.route('/total-budget')
@login_required
def total_budget():
    conn = get_db()
    cursor = conn.cursor()

    income_types = ['bank', 'land', 'home', 'wedding', 'other']
    income_labels = {
        'bank': 'ធនាគារ',
        'land': 'ការងារដីធ្លី',
        'home': 'ការងារផ្ទះ',
        'wedding': 'អាពាហ៍ពិពាហ៍',
        'other': 'ចូលផ្សេងៗ'
    }
    income_colors = {
        'bank': '#4CAF50',
        'land': '#2196F3',
        'home': '#FF9800',
        'wedding': '#E91E63',
        'other': '#9C27B0'
    }

    income_data = {}
    total_income_all = 0

    for t in income_types:
        cursor.execute('SELECT SUM(total_price) FROM income_records WHERE type = ?', (t,))
        amount = cursor.fetchone()[0] or 0
        income_data[t] = amount
        total_income_all += amount

    expense_types = ['report', 'decision', 'warrant', 'financial', 'other']
    expense_labels = {
        'report': 'របាយការណ៍',
        'decision': 'សេចក្តីសម្រេច',
        'warrant': 'របក ដីការ',
        'financial': 'បៀរវត្ស/ហិរញ្ញវត្ថុ',
        'other': 'ឯកសារចេញផ្សេងៗ'
    }
    expense_colors = {
        'report': '#FF6B6B',
        'decision': '#FF9F43',
        'warrant': '#FECA57',
        'financial': '#48DBFB',
        'other': '#A29BFE'
    }

    expense_data = {}
    total_expense_all = 0

    for t in expense_types:
        cursor.execute('SELECT SUM(total_price) FROM expense_records WHERE type = ?', (t,))
        amount = cursor.fetchone()[0] or 0
        expense_data[t] = amount
        total_expense_all += amount

    conn.close()

    chart_data = {
        'income_labels': list(income_labels.values()),
        'income_values': [income_data.get(t, 0) for t in income_types],
        'income_colors': [income_colors[t] for t in income_types],
        'expense_labels': list(expense_labels.values()),
        'expense_values': [expense_data.get(t, 0) for t in expense_types],
        'expense_colors': [expense_colors[t] for t in expense_types]
    }

    return render_template('total_budget.html',
                         username=session['username'],
                         income_data=income_data,
                         income_labels=income_labels,
                         income_colors=income_colors,
                         expense_data=expense_data,
                         expense_labels=expense_labels,
                         expense_colors=expense_colors,
                         total_income=total_income_all,
                         total_expense=total_expense_all,
                         balance=total_income_all - total_expense_all,
                         chart_data=chart_data,
                         permissions=get_user_permissions())


# ============================================
# ROUTE: ពត៌មាន (Info)
# ============================================
@app.route('/info')
@login_required
def info():
    user_permissions = get_user_permissions()
    return render_template('info.html',
                         username=session['username'],
                         permissions=user_permissions)


# ============================================
# ROUTE: ការកំណត់ (Settings)
# ============================================
@app.route('/settings')
@login_required
def settings():
    return render_template('settings.html',
                         username=session['username'],
                         permissions=get_user_permissions())


# ============================================
# ROUTE: គ្រប់គ្រងបុគ្គលិក (Employees)
# ============================================
@app.route('/employees')
@login_required
def employees():
    user_permissions = get_user_permissions()
    if 'all' not in user_permissions and 'employees_view' not in user_permissions:
        flash('អ្នកមិនមានសិទ្ធិចូលប្រើទំព័រនេះទេ!', 'error')
        return redirect(url_for('dashboard'))

    return render_template('employees.html',
                         username=session['username'],
                         permissions=user_permissions)


# ===== API: GET Info Documents =====
@app.route('/api/info-documents', methods=['GET'])
@login_required
def api_get_info_documents():
    if not has_permission('info_view'):
        return jsonify([])

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM info_documents ORDER BY category')
    docs = cursor.fetchall()
    conn.close()

    result = {}
    for row in docs:
        result[row['category']] = {
            'id': row['id'],
            'file_path': row['file_path'],
            'file_name': row['file_name'],
            'uploaded_at': row['uploaded_at'],
            'uploaded_by': row['uploaded_by']
        }

    return jsonify(result)


# ===== API: UPLOAD Info Document =====
@app.route('/api/info-documents', methods=['POST'])
@login_required
def api_upload_info_document():
    if not has_permission('info_upload'):
        return jsonify({'success': False, 'error': 'អ្នកមិនមានសិទ្ធិ!'})

    category = request.form.get('category')
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'គ្មានឯកសារត្រូវបានផ្ទុកឡើង!'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'គ្មានឯកសារត្រូវបានជ្រើសរើស!'})

    allowed_extensions = ['pdf', 'png', 'jpg', 'jpeg', 'gif']
    file_ext = file.filename.rsplit('.', 1)[1].lower()
    if file_ext not in allowed_extensions:
        return jsonify({'success': False, 'error': 'ប្រភេទឯកសារមិនត្រូវបានអនុញ្ញាត! (PDF, PNG, JPG, JPEG, GIF)'})

    filename = f"info_{category}_{get_cambodia_time().strftime('%Y%m%d_%H%M%S')}.{file_ext}"
    file_path = os.path.join('uploads', filename)

    os.makedirs('uploads', exist_ok=True)
    file.save(file_path)

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT file_path FROM info_documents WHERE category = ?', (category,))
    old_doc = cursor.fetchone()
    if old_doc and old_doc['file_path']:
        delete_file_if_exists(old_doc['file_path'])

    cursor.execute('DELETE FROM info_documents WHERE category = ?', (category,))

    cursor.execute('''
        INSERT INTO info_documents (category, file_path, file_name, uploaded_at, uploaded_by)
        VALUES (?, ?, ?, ?, ?)
    ''', (category, file_path, file.filename, get_cambodia_time_str(), session['username']))
    conn.commit()
    conn.close()

    return jsonify({'success': True, 'file_path': file_path, 'file_name': file.filename})


# ===== API: DELETE Info Document =====
@app.route('/api/info-documents/<category>', methods=['DELETE'])
@login_required
def api_delete_info_document(category):
    if not has_permission('info_upload'):
        return jsonify({'success': False, 'error': 'អ្នកមិនមានសិទ្ធិ!'})

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT file_path FROM info_documents WHERE category = ?', (category,))
    old_doc = cursor.fetchone()
    if old_doc and old_doc['file_path']:
        delete_file_if_exists(old_doc['file_path'])

    cursor.execute('DELETE FROM info_documents WHERE category = ?', (category,))
    conn.commit()
    conn.close()

    log_audit(
        action='លុបឯកសារពត៌មាន',
        module='ពត៌មាន',
        details=f"បានលុបឯកសារប្រភេទ {category}"
    )

    return jsonify({'success': True})


# ============================================
# ROUTE: បម្រើឯកសារពីថត uploads
# ============================================
@app.route('/uploads/<path:filename>')
@login_required
def uploaded_file(filename):
    try:
        file_path = os.path.join('uploads', filename)
        if not os.path.exists(file_path):
            return jsonify({'error': 'ឯកសារមិនមាន'}), 404
        return send_from_directory('uploads', filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 404


# ===== API: GET Employees =====
@app.route('/api/employees', methods=['GET'])
@login_required
def api_get_employees():
    if not has_permission('employees_view'):
        return jsonify([])

    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM employees ORDER BY id DESC')
        employees = cursor.fetchall()
        conn.close()

        result = []
        for row in employees:
            result.append({
                'id': row['id'],
                'photo': row['photo'] or '',
                'name': row['name'],
                'gender': row['gender'],
                'birth_date': row['birth_date'],
                'birth_place': row['birth_place'],
                'id_card': row['id_card'],
                'role': row['role'],
                'salary': row['salary'],
                'start_date': row['start_date'],
                'start_file': row['start_file'] or '',
                'end_date': row['end_date'] or '',
                'end_file': row['end_file'] or '',
                'file_path': row['file_path'] or '',
                'created_at': row['created_at']
            })

        return jsonify(result)
    except Exception as e:
        print(f"Error in api_get_employees: {e}")
        return jsonify([])


# ===== API: ADD Employee =====
@app.route('/api/employees', methods=['POST'])
@login_required
def api_add_employee():
    if not has_permission('employees_add'):
        return jsonify({'success': False, 'error': 'អ្នកមិនមានសិទ្ធិបន្ថែមបុគ្គលិកទេ!'})

    data = request.json

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO employees
        (photo, name, gender, birth_date, birth_place, id_card, role, salary, start_date, start_file, end_date, end_file, created_at, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data.get('photo', ''),
        data.get('name'),
        data.get('gender'),
        data.get('birth_date'),
        data.get('birth_place'),
        data.get('id_card'),
        data.get('role'),
        data.get('salary'),
        data.get('start_date'),
        data.get('start_file', ''),
        data.get('end_date', ''),
        data.get('end_file', ''),
        get_cambodia_time_str(),
        session['username']
    ))
    conn.commit()
    new_id = cursor.lastrowid
    conn.close()

    log_audit(
        action='បន្ថែមបុគ្គលិក',
        module='បុគ្គលិក',
        record_id=new_id,
        details=f"បានបន្ថែមបុគ្គលិក {data.get('name')}"
    )

    return jsonify({'success': True, 'id': new_id})


# ===== API: UPDATE Employee =====
@app.route('/api/employees/<int:employee_id>', methods=['PUT'])
@login_required
def api_update_employee(employee_id):
    if not has_permission('employees_edit'):
        return jsonify({'success': False, 'error': 'អ្នកមិនមានសិទ្ធិកែប្រែបុគ្គលិកទេ!'})

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT photo, start_file, end_file FROM employees WHERE id = ?', (employee_id,))
    old_data = cursor.fetchone()

    data = request.json

    if data.get('photo') and old_data and old_data['photo'] and old_data['photo'] != data.get('photo'):
        delete_file_if_exists(old_data['photo'])

    if data.get('start_file') and old_data and old_data['start_file'] and old_data['start_file'] != data.get('start_file'):
        delete_file_if_exists(old_data['start_file'])

    if data.get('end_file') and old_data and old_data['end_file'] and old_data['end_file'] != data.get('end_file'):
        delete_file_if_exists(old_data['end_file'])

    cursor.execute('''
        UPDATE employees
        SET photo = ?, name = ?, gender = ?, birth_date = ?, birth_place = ?,
            id_card = ?, role = ?, salary = ?, start_date = ?, start_file = ?,
            end_date = ?, end_file = ?
        WHERE id = ?
    ''', (
        data.get('photo', ''),
        data.get('name'),
        data.get('gender'),
        data.get('birth_date'),
        data.get('birth_place'),
        data.get('id_card'),
        data.get('role'),
        data.get('salary'),
        data.get('start_date'),
        data.get('start_file', ''),
        data.get('end_date', ''),
        data.get('end_file', ''),
        employee_id
    ))
    conn.commit()
    conn.close()

    log_audit(
        action='កែប្រែបុគ្គលិក',
        module='បុគ្គលិក',
        record_id=employee_id,
        details=f"បានកែប្រែបុគ្គលិក ID {employee_id}"
    )

    return jsonify({'success': True})


# ===== API: DELETE Employee =====
@app.route('/api/employees/<int:employee_id>', methods=['DELETE'])
@login_required
def api_delete_employee(employee_id):
    if not has_permission('employees_delete'):
        return jsonify({'success': False, 'error': 'អ្នកមិនមានសិទ្ធិលុបបុគ្គលិកទេ!'})

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT photo, start_file, end_file FROM employees WHERE id = ?', (employee_id,))
    files = cursor.fetchone()

    if files:
        delete_file_if_exists(files['photo'])
        delete_file_if_exists(files['start_file'])
        delete_file_if_exists(files['end_file'])

    cursor.execute('DELETE FROM employees WHERE id = ?', (employee_id,))
    conn.commit()
    conn.close()

    log_audit(
        action='លុបបុគ្គលិក',
        module='បុគ្គលិក',
        record_id=employee_id,
        details=f"បានលុបបុគ្គលិក ID {employee_id}"
    )

    return jsonify({'success': True})


# ===== API: UPLOAD File =====
@app.route('/api/upload', methods=['POST'])
@login_required
def api_upload_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'គ្មានឯកសារត្រូវបានផ្ទុកឡើង!'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'គ្មានឯកសារត្រូវបានជ្រើសរើស!'})

    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}
    file_extension = file.filename.rsplit('.', 1)[1].lower()
    if file_extension not in allowed_extensions:
        return jsonify({'success': False, 'error': 'ប្រភេទឯកសារមិនត្រូវបានអនុញ្ញាត! (អនុញ្ញាតតែ: png, jpg, jpeg, gif, pdf)'})

    filename = f"{get_cambodia_time().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
    file_path = os.path.join('uploads', filename)

    os.makedirs('uploads', exist_ok=True)
    file.save(file_path)

    return jsonify({'success': True, 'file_path': file_path, 'filename': filename})


# ============================================
# API: USERS
# ============================================
@app.route('/api/users', methods=['GET'])
@login_required
def api_get_users():
    if not has_permission('settings_users'):
        return jsonify([])

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT u.*, r.name as role_name
        FROM users u
        LEFT JOIN roles r ON u.role_id = r.id
        ORDER BY u.id
    ''')
    users = cursor.fetchall()
    conn.close()

    result = []
    for row in users:
        result.append({
            'id': row['id'],
            'username': row['username'],
            'full_name': row['full_name'] or '',
            'phone': row['phone'] or '',
            'role_id': row['role_id'],
            'role_name': row['role_name'] or 'No Role',
            'is_active': row['is_active'],
            'created_at': row['created_at']
        })

    return jsonify(result)


@app.route('/api/users', methods=['POST'])
@login_required
def api_add_user():
    if not has_permission('settings_users'):
        return jsonify({'success': False, 'error': 'អ្នកមិនមានសិទ្ធិ!'})

    data = request.json

    phone = data.get('phone', '').strip()
    if not phone:
        return jsonify({'success': False, 'error': 'សូមបញ្ចូលលេខទូរស័ព្ទរបស់អ្នកប្រើប្រាស់!'})

    phone_pattern = r'^[0-9]{9,10}$'
    if not re.match(phone_pattern, phone):
        return jsonify({'success': False, 'error': 'លេខទូរស័ព្ទមិនត្រឹមត្រូវ! (ត្រូវមាន 9-10 ខ្ទង់)'})

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT COUNT(*) FROM users WHERE username = ?', (data.get('username'),))
    if cursor.fetchone()[0] > 0:
        conn.close()
        return jsonify({'success': False, 'error': 'ឈ្មោះអ្នកប្រើប្រាស់មានរួចហើយ!'})

    cursor.execute('SELECT COUNT(*) FROM users WHERE phone = ?', (phone,))
    if cursor.fetchone()[0] > 0:
        conn.close()
        return jsonify({'success': False, 'error': 'លេខទូរស័ព្ទនេះត្រូវបានប្រើរួចហើយ!'})

    password = data.get('password', 'password123')

    cursor.execute('''
        INSERT INTO users (username, password, full_name, phone, role_id, is_active, created_at, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data.get('username'),
        password,
        data.get('full_name', ''),
        phone,
        data.get('role_id', 2),
        data.get('is_active', 1),
        get_cambodia_time_str(),
        session['username']
    ))
    conn.commit()
    new_id = cursor.lastrowid
    conn.close()

    log_audit(
        action='បន្ថែមអ្នកប្រើប្រាស់',
        module='ការកំណត់',
        record_id=new_id,
        details=f"បានបន្ថែមអ្នកប្រើប្រាស់ {data.get('username')} (Phone: {phone})"
    )

    return jsonify({'success': True, 'id': new_id})


@app.route('/api/users/<int:user_id>', methods=['PUT'])
@login_required
def api_update_user(user_id):
    if not has_permission('settings_users'):
        return jsonify({'success': False, 'error': 'អ្នកមិនមានសិទ្ធិ!'})

    data = request.json

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT username FROM users WHERE id = ?', (user_id,))
    old_user = cursor.fetchone()
    old_username = old_user['username'] if old_user else 'unknown'

    phone = data.get('phone', '').strip()
    if phone:
        phone_pattern = r'^[0-9]{9,10}$'
        if not re.match(phone_pattern, phone):
            conn.close()
            return jsonify({'success': False, 'error': 'លេខទូរស័ព្ទមិនត្រឹមត្រូវ!'})

    if data.get('password'):
        cursor.execute('''
            UPDATE users
            SET full_name = ?, phone = ?, role_id = ?, is_active = ?, password = ?
            WHERE id = ?
        ''', (
            data.get('full_name', ''),
            phone,
            data.get('role_id'),
            data.get('is_active', 1),
            data.get('password'),
            user_id
        ))
    else:
        cursor.execute('''
            UPDATE users
            SET full_name = ?, phone = ?, role_id = ?, is_active = ?
            WHERE id = ?
        ''', (
            data.get('full_name', ''),
            phone,
            data.get('role_id'),
            data.get('is_active', 1),
            user_id
        ))

    conn.commit()
    conn.close()

    log_audit(
        action='កែប្រែអ្នកប្រើប្រាស់',
        module='ការកំណត់',
        record_id=user_id,
        details=f"បានកែប្រែអ្នកប្រើប្រាស់ {old_username}"
    )

    return jsonify({'success': True})


@app.route('/api/users/<int:user_id>/reset-password', methods=['POST'])
@login_required
def api_reset_password(user_id):
    if not has_permission('settings_users'):
        return jsonify({'success': False, 'error': 'អ្នកមិនមានសិទ្ធិ!'})

    data = request.json
    new_password = data.get('new_password', 'password123')

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT username FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()
    username = user['username'] if user else 'unknown'

    cursor.execute('UPDATE users SET password = ? WHERE id = ?', (new_password, user_id))
    conn.commit()
    conn.close()

    log_audit(
        action='កំណត់ពាក្យសម្ងាត់ឡើងវិញ',
        module='ការកំណត់',
        record_id=user_id,
        details=f"បានកំណត់ពាក្យសម្ងាត់ឡើងវិញសម្រាប់អ្នកប្រើប្រាស់ {username}"
    )

    return jsonify({'success': True})


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
def api_delete_user(user_id):
    if not has_permission('settings_users'):
        return jsonify({'success': False, 'error': 'អ្នកមិនមានសិទ្ធិ!'})

    if user_id == session.get('user_id'):
        return jsonify({'success': False, 'error': 'មិនអាចលុបខ្លួនឯងបានទេ!'})

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT username FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()
    username = user['username'] if user else 'unknown'

    cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
    conn.commit()
    conn.close()

    log_audit(
        action='លុបអ្នកប្រើប្រាស់',
        module='ការកំណត់',
        record_id=user_id,
        details=f"បានលុបអ្នកប្រើប្រាស់ {username}"
    )

    return jsonify({'success': True})


# ============================================
# API: ROLES
# ============================================
@app.route('/api/roles', methods=['GET'])
@login_required
def api_get_roles():
    if not has_permission('settings_roles'):
        return jsonify([])

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM roles ORDER BY id')
    roles = cursor.fetchall()
    conn.close()

    result = []
    for row in roles:
        result.append({
            'id': row['id'],
            'name': row['name'],
            'description': row['description'] or '',
            'permissions': row['permissions'] or ''
        })

    return jsonify(result)


@app.route('/api/roles', methods=['POST'])
@login_required
def api_add_role():
    if not has_permission('settings_roles'):
        return jsonify({'success': False, 'error': 'អ្នកមិនមានសិទ្ធិ!'})

    data = request.json

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO roles (name, description, permissions, created_at, created_by)
        VALUES (?, ?, ?, ?, ?)
    ''', (
        data.get('name'),
        data.get('description', ''),
        data.get('permissions', ''),
        get_cambodia_time_str(),
        session['username']
    ))
    conn.commit()
    new_id = cursor.lastrowid
    conn.close()

    log_audit(
        action='បន្ថែមតួនាទី',
        module='ការកំណត់',
        record_id=new_id,
        details=f"បានបន្ថែមតួនាទី {data.get('name')}"
    )

    return jsonify({'success': True, 'id': new_id})


@app.route('/api/roles/<int:role_id>', methods=['PUT'])
@login_required
def api_update_role(role_id):
    if not has_permission('settings_roles'):
        return jsonify({'success': False, 'error': 'អ្នកមិនមានសិទ្ធិ!'})

    data = request.json

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT name FROM roles WHERE id = ?', (role_id,))
    old_role = cursor.fetchone()
    old_name = old_role['name'] if old_role else 'unknown'

    cursor.execute('''
        UPDATE roles
        SET name = ?, description = ?, permissions = ?
        WHERE id = ?
    ''', (
        data.get('name'),
        data.get('description', ''),
        data.get('permissions', ''),
        role_id
    ))
    conn.commit()
    conn.close()

    log_audit(
        action='កែប្រែតួនាទី',
        module='ការកំណត់',
        record_id=role_id,
        details=f"បានកែប្រែតួនាទី {old_name}"
    )

    return jsonify({'success': True})


@app.route('/api/roles/<int:role_id>', methods=['DELETE'])
@login_required
def api_delete_role(role_id):
    if not has_permission('settings_roles'):
        return jsonify({'success': False, 'error': 'អ្នកមិនមានសិទ្ធិ!'})

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT name FROM roles WHERE id = ?', (role_id,))
    role = cursor.fetchone()
    role_name = role['name'] if role else 'unknown'

    cursor.execute('DELETE FROM roles WHERE id = ?', (role_id,))
    conn.commit()
    conn.close()

    log_audit(
        action='លុបតួនាទី',
        module='ការកំណត់',
        record_id=role_id,
        details=f"បានលុបតួនាទី {role_name}"
    )

    return jsonify({'success': True})


# ============================================
# API: AUDIT LOGS
# ============================================
@app.route('/api/audit-logs', methods=['GET'])
@login_required
def api_get_audit_logs():
    if not has_permission('settings_audit'):
        return jsonify([])

    search = request.args.get('search', '').strip()

    conn = get_db()
    cursor = conn.cursor()

    if search:
        query = '''
            SELECT * FROM audit_logs
            WHERE username LIKE ? OR action LIKE ? OR module LIKE ? OR details LIKE ?
            ORDER BY id DESC LIMIT 100
        '''
        search_param = f'%{search}%'
        cursor.execute(query, (search_param, search_param, search_param, search_param))
    else:
        cursor.execute('SELECT * FROM audit_logs ORDER BY id DESC LIMIT 100')

    logs = cursor.fetchall()
    conn.close()

    result = []
    for row in logs:
        result.append({
            'id': row['id'],
            'user_id': row['user_id'],
            'username': row['username'],
            'action': row['action'],
            'module': row['module'],
            'record_id': row['record_id'],
            'details': row['details'] or '',
            'created_at': row['created_at']
        })

    return jsonify(result)


# ============================================
# API: REQUEST OTP
# ============================================
@app.route('/api/request-otp', methods=['POST'])
def api_request_otp():
    """ស្នើសុំ OTP តាមទូរស័ព្ទ"""
    data = request.json
    username = data.get('username', '').strip()
    phone = data.get('phone', '').strip()

    if not username or not phone:
        return jsonify({'success': False, 'error': 'សូមបំពេញព័ត៌មានទាំងអស់!'})

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT id, username, phone FROM users
        WHERE username = ? AND phone = ? AND is_active = 1
    ''', (username, phone))

    user = cursor.fetchone()

    if not user:
        conn.close()
        return jsonify({'success': False, 'error': 'ឈ្មោះអ្នកប្រើប្រាស់ ឬ លេខទូរស័ព្ទមិនត្រឹមត្រូវ!'})

    # ===== បង្កើត OTP =====
    otp_code = generate_otp(6)
    expires_at = (get_cambodia_time() + timedelta(minutes=5)).strftime('%Y-%m-%d %H:%M:%S')

    # ===== លុប OTP ចាស់ =====
    cursor.execute('DELETE FROM otp_codes WHERE user_id = ?', (user['id'],))

    # ===== រក្សាទុក OTP =====
    cursor.execute('''
        INSERT INTO otp_codes (user_id, phone, otp_code, expires_at, created_at)
        VALUES (?, ?, ?, ?, ?)
    ''', (user['id'], phone, otp_code, expires_at, get_cambodia_time_str()))
    conn.commit()
    conn.close()

    # ===== ផ្ញើ OTP តាម SMS =====
    message = f'OTP របស់អ្នកគឺ: {otp_code} (មានសុពលភាព 5 នាទី)'
    sms_sent = send_sms(phone, message)

    if sms_sent:
        return jsonify({'success': True, 'message': 'បានផ្ញើ OTP ទៅកាន់ទូរស័ព្ទរបស់អ្នក!'})
    else:
        # ===== Fallback: បង្ហាញ OTP ក្នុង Response =====
        return jsonify({
            'success': True,
            'message': f'OTP: {otp_code} (សម្រាប់សាកល្បង - មិនអាចផ្ញើ SMS បាន)'
        })

# ============================================
# API: VERIFY OTP
# ============================================
@app.route('/api/verify-otp', methods=['POST'])
def api_verify_otp():
    """ផ្ទៀងផ្ទាត់ OTP និងបង្កើតពាក្យសម្ងាត់ថ្មី"""
    data = request.json
    username = data.get('username', '').strip()
    phone = data.get('phone', '').strip()
    otp_code = data.get('otp', '').strip()

    if not username or not phone or not otp_code:
        return jsonify({'success': False, 'error': 'សូមបំពេញព័ត៌មានទាំងអស់!'})

    if len(otp_code) != 6:
        return jsonify({'success': False, 'error': 'OTP ត្រូវមាន 6 ខ្ទង់!'})

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT u.id, u.username
        FROM otp_codes o
        JOIN users u ON o.user_id = u.id
        WHERE u.username = ? AND o.phone = ?
        AND o.otp_code = ?
        AND o.is_used = 0
        AND o.expires_at > ?
        ORDER BY o.id DESC LIMIT 1
    ''', (username, phone, otp_code, get_cambodia_time_str()))

    otp = cursor.fetchone()

    if not otp:
        conn.close()
        return jsonify({'success': False, 'error': 'OTP មិនត្រឹមត្រូវ ឬ ផុតកំណត់!'})

    # ===== សម្គាល់ OTP ថាបានប្រើ =====
    cursor.execute('UPDATE otp_codes SET is_used = 1 WHERE user_id = ?', (otp['id'],))

    # ===== បង្កើតពាក្យសម្ងាត់ថ្មី =====
    new_password = generate_random_password(10)
    cursor.execute('UPDATE users SET password = ? WHERE id = ?', (new_password, otp['id']))
    conn.commit()
    conn.close()

    log_audit(
        action='ស្នើសុំពាក្យសម្ងាត់ថ្មី (OTP)',
        module='ប្រព័ន្ធ',
        record_id=otp['id'],
        details=f"អ្នកប្រើប្រាស់ {username} បានស្នើសុំពាក្យសម្ងាត់ថ្មីតាម OTP"
    )

    return jsonify({
        'success': True,
        'message': f'បានកំណត់ពាក្យសម្ងាត់ថ្មីរួចរាល់!',
        'new_password': new_password  # ← បន្ថែមនេះ
    })


# ============================================
# FAVICON
# ============================================
@app.route('/favicon.ico')
def favicon():
    svg = '''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 100 100">
        <rect width="100" height="100" rx="20" fill="#1a1a2e"/>
        <text x="50" y="72" font-size="60" text-anchor="middle" font-family="Arial" fill="#FFD700">📊</text>
    </svg>'''
    return svg, 200, {
        'Content-Type': 'image/svg+xml',
        'Cache-Control': 'public, max-age=86400'
    }


# ============================================
# MANIFEST.JSON
# ============================================
@app.route('/manifest.json')
def manifest():
    manifest_data = {
        "name": "ប្រព័ន្ធគ្រប់គ្រងរដ្ឋបាល",
        "short_name": "គ្រប់គ្រងរដ្ឋបាល",
        "description": "ប្រព័ន្ធគ្រប់គ្រងរដ្ឋបាល - ដំណើរការបានទាំង Online និង Offline",
        "start_url": "/dashboard",
        "display": "standalone",
        "background_color": "#0a0a0f",
        "theme_color": "#FFD700",
        "orientation": "portrait",
        "scope": "/",
        "icons": [
            {
                "src": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 512 512'%3E%3Crect width='512' height='512' rx='64' fill='%231a1a2e'/%3E%3Ctext x='256' y='380' font-size='280' text-anchor='middle' font-family='Arial' fill='%23FFD700'%3E📊%3C/text%3E%3C/svg%3E",
                "sizes": "512x512",
                "type": "image/svg+xml",
                "purpose": "any maskable"
            },
            {
                "src": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 192 192'%3E%3Crect width='192' height='192' rx='24' fill='%231a1a2e'/%3E%3Ctext x='96' y='142' font-size='105' text-anchor='middle' font-family='Arial' fill='%23FFD700'%3E📊%3C/text%3E%3C/svg%3E",
                "sizes": "192x192",
                "type": "image/svg+xml",
                "purpose": "any maskable"
            }
        ]
    }
    return jsonify(manifest_data)


# ============================================
# SERVICE WORKER
# ============================================
@app.route('/service-worker.js')
def service_worker():
    js = '''// Service Worker - Offline Support
const CACHE_NAME = 'admin-system-v1';
const STATIC_FILES = [
    '/',
    '/dashboard',
    '/income',
    '/expense',
    '/income-budget',
    '/expense-budget',
    '/total-budget',
    '/customers',
    '/info',
    '/employees',
    '/settings',
    '/offline'
];

self.addEventListener('install', event => {
    event.waitUntil(
        caches.open(CACHE_NAME)
            .then(cache => cache.addAll(STATIC_FILES))
            .then(() => self.skipWaiting())
    );
});

self.addEventListener('activate', event => {
    event.waitUntil(
        caches.keys().then(cacheNames => {
            return Promise.all(
                cacheNames.map(name => {
                    if (name !== CACHE_NAME) {
                        return caches.delete(name);
                    }
                })
            );
        }).then(() => self.clients.claim())
    );
});

self.addEventListener('fetch', event => {
    event.respondWith(
        fetch(event.request)
            .then(response => {
                const clone = response.clone();
                caches.open(CACHE_NAME).then(cache => {
                    cache.put(event.request, clone);
                });
                return response;
            })
            .catch(() => {
                return caches.match(event.request);
            })
    );
});

console.log('Service Worker loaded successfully!');'''
    return js, 200, {
        'Content-Type': 'application/javascript',
        'Cache-Control': 'public, max-age=86400'
    }


# ============================================
# LOGO
# ============================================
@app.route('/static/Logo.png')
def serve_logo():
    try:
        return send_file('static/Logo.png')
    except:
        try:
            from PIL import Image, ImageDraw, ImageFont
            import io

            img = Image.new('RGB', (200, 200), color='#1a1a2e')
            draw = ImageDraw.Draw(img)
            draw.rounded_rectangle([0, 0, 200, 200], radius=25, fill='#1a1a2e')

            try:
                font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 100)
            except:
                font = ImageFont.load_default()

            draw.text((100, 100), "📊", fill='#FFD700', anchor="mm", font=font)

            output = io.BytesIO()
            img.save(output, format='PNG')
            output.seek(0)
            return send_file(output, mimetype='image/png')
        except:
            return "Logo not available", 404


# ============================================
# API: CHANGE PASSWORD
# ============================================
@app.route('/api/change-password', methods=['POST'])
@login_required
def api_change_password():
    data = request.json
    current_password = data.get('current_password')
    new_password = data.get('new_password')

    if not current_password or not new_password:
        return jsonify({'success': False, 'error': 'សូមបំពេញព័ត៌មានទាំងអស់!'})

    if len(new_password) < 6:
        return jsonify({'success': False, 'error': 'ពាក្យសម្ងាត់ថ្មីត្រូវមានយ៉ាងហោចណាស់ 6 តួ!'})

    conn = get_db()
    cursor = conn.cursor()

    user_id = session.get('user_id')
    cursor.execute('SELECT password FROM users WHERE id = ?', (user_id,))
    row = cursor.fetchone()

    if not row:
        conn.close()
        return jsonify({'success': False, 'error': 'មិនឃើញអ្នកប្រើប្រាស់!'})

    if row['password'] != current_password:
        conn.close()
        return jsonify({'success': False, 'error': 'ពាក្យសម្ងាត់បច្ចុប្បន្នមិនត្រឹមត្រូវ!'})

    cursor.execute('UPDATE users SET password = ? WHERE id = ?', (new_password, user_id))
    conn.commit()
    conn.close()

    log_audit(
        action='ប្តូរពាក្យសម្ងាត់',
        module='ប្រព័ន្ធ',
        details=f"អ្នកប្រើប្រាស់ {session.get('username')} បានប្តូរពាក្យសម្ងាត់"
    )

    return jsonify({'success': True, 'message': 'បានប្តូរពាក្យសម្ងាត់ដោយជោគជ័យ!'})

# ============================================
# TWILIO SMS VIA REQUESTS (NO LIBRARY)
# ============================================
import requests
import base64

TWILIO_CONFIG = {
    'account_sid': 'ACb8ac73cba4ae97d31d882f8e0279b0ea',
    'auth_token': 'b043b709931491a37dcf8bb733de778b',
    'from_phone': '+14845595918'
}

def send_sms(phone, message):
    """ផ្ញើ SMS តាមរយៈ Twilio API"""
    try:
        if not phone.startswith('+'):
            if phone.startswith('0'):
                phone = '+855' + phone[1:]
            else:
                phone = '+855' + phone

        url = f"https://api.twilio.com/2010-04-01/Accounts/{TWILIO_CONFIG['account_sid']}/Messages.json"
        auth_string = f"{TWILIO_CONFIG['account_sid']}:{TWILIO_CONFIG['auth_token']}"
        auth_b64 = base64.b64encode(auth_string.encode('utf-8')).decode('utf-8')

        data = {
            'From': TWILIO_CONFIG['from_phone'],
            'To': phone,
            'Body': message
        }
        headers = {
            'Authorization': f'Basic {auth_b64}',
            'Content-Type': 'application/x-www-form-urlencoded'
        }

        response = requests.post(url, data=data, headers=headers)
        return response.status_code == 201
    except Exception as e:
        print(f"❌ Error sending SMS: {e}")
        return False


# ============================================
# TWILIO SMS VIA REQUESTS (NO LIBRARY)
# ============================================
import requests
import base64

TWILIO_CONFIG = {
    'account_sid': 'ACb8ac73cba4ae97d31d882f8e0279b0ea',
    'auth_token': 'b043b709931491a37dcf8bb733de778b',
    'from_phone': '+14845595918'
}

def send_sms(phone, message):
    """ផ្ញើ SMS តាមរយៈ Twilio API (ដោយមិនប្រើ Library)"""
    try:
        # ===== កែទម្រង់លេខទូរស័ព្ទ =====
        if not phone.startswith('+'):
            if phone.startswith('0'):
                phone = '+855' + phone[1:]
            else:
                phone = '+855' + phone

        url = f"https://api.twilio.com/2010-04-01/Accounts/{TWILIO_CONFIG['account_sid']}/Messages.json"

        # ===== បង្កើត Authorization =====
        auth_string = f"{TWILIO_CONFIG['account_sid']}:{TWILIO_CONFIG['auth_token']}"
        auth_bytes = auth_string.encode('utf-8')
        auth_b64 = base64.b64encode(auth_bytes).decode('utf-8')

        data = {
            'From': TWILIO_CONFIG['from_phone'],
            'To': phone,
            'Body': message
        }

        headers = {
            'Authorization': f'Basic {auth_b64}',
            'Content-Type': 'application/x-www-form-urlencoded'
        }

        print(f"📤 Sending SMS to {phone}...")
        response = requests.post(url, data=data, headers=headers)

        if response.status_code == 201:
            print(f"✅ SMS sent to {phone}")
            return True
        else:
            print(f"❌ SMS failed: {response.status_code} - {response.text}")
            return False
    except Exception as e:
        print(f"❌ Error sending SMS: {e}")
        return False

# ============================================
# INIT DATABASE
# ============================================
init_db()


# ============================================
# RUN APP
# ============================================
if __name__ == '__main__':
    app.run(debug=True)
